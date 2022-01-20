#Excelファイルの新規作成(カテゴリ分け前のデータ作成用)
#今回はExcelデータ作成時pandasを用いずにエクセルの元データを作成してみる
import openpyxl as excel
import random

wbname = "test.xlsx"
wb = excel.Workbook()
ws = wb["Sheet"]
ws["A1"] = "都道府県"
ws["B1"] = "名前"
ws["C1"] = "年齢"

#都道府県リスト
prefecture = ["北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県",
              "福島県", "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県",
              "東京都", "神奈川県", "新潟県", "富山県", "石川県", "福井県",
              "山梨県", "長野県", "岐阜県", "静岡県", "愛知県", "三重県",
              "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
              "鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県",
              "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
              "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"]

#名前リスト(苗字)
last_name = ["佐藤", "鈴木", "高橋", "田中", "伊藤", "渡辺",
             "山本", "中村", "小林", "加藤"]

#名前リスト(名前)
first_name = ["蓮", "陽翔", "蒼", "湊", "樹",
              "紬", "陽葵", "凛", "澪", "芽依"]

#何件のデータを作成するかをユーザに確認
#数値が入るまで繰り返す
while(1):
  count = input("何件のデータを作成しますか？>>>")
  if count.isdecimal():
    count = int(count)
    break
  else:
    print("数値を入力してください")

#count分だけデータを作成
#１つのデータは都道府県、名前、年齢
data = []
for i in range(count):
  data.append({"都道府県": random.choice(prefecture),
               "名前": random.choice(last_name) + " " + random.choice(first_name),
               "年齢": random.randint(0, 100)})
  
  ws.cell(i + 2, 1, value = data[i]["都道府県"])
  ws.cell(i + 2, 2, value = data[i]["名前"])
  ws.cell(i + 2, 3, value = data[i]["年齢"])
  
#保存
wb.save("test.xlsx")
